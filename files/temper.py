import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

groups = ["F-7324", "DII-23"]

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Schedule Template"

# Define column widths
# A and B are fixed at 10
column_widths = {'A': 10, 'B': 10}

# Dynamically set widths for group columns (C onward)
# Pattern: 25, 15, 20, 15 (repeats every 4 columns)
columns_per_group = 4
group_column_widths = [25, 15, 20, 15]  # Widths for Fan, Dars turi, O'qituvchi, Xona
for idx in range(len(groups) * columns_per_group):
    col_letter = get_column_letter(3 + idx)  # Start from column C (3rd column)
    width_index = idx % 4  # Repeat the pattern every 4 columns
    column_widths[col_letter] = group_column_widths[width_index]

# Apply column widths
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Define styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
center_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Write the template for columns A and B
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
ws.cell(row=3, column=1).value = "Guruhlar ->"
ws.cell(row=3, column=1).font = header_font
ws.cell(row=3, column=1).fill = header_fill
ws.cell(row=3, column=1).alignment = center_alignment

ws.cell(row=4, column=1).value = "kunlar"
ws.cell(row=4, column=1).font = header_font
ws.cell(row=4, column=1).fill = header_fill
ws.cell(row=4, column=1).alignment = center_alignment

ws.cell(row=4, column=2).value = "paralar"
ws.cell(row=4, column=2).font = header_font
ws.cell(row=4, column=2).fill = header_fill
ws.cell(row=4, column=2).alignment = center_alignment

days = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba"]
for idx, day in enumerate(days):
    start_row = 5 + idx * 6  # Start from row 5
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 5, end_column=1)
    day_cell = ws.cell(row=start_row, column=1)
    day_cell.value = day
    day_cell.font = header_font
    day_cell.fill = header_fill
    day_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)  # Rotate text vertically

    for para in range(1, 7):
        ws.cell(row=start_row + para - 1, column=2).value = para
        ws.cell(row=start_row + para - 1, column=2).alignment = center_alignment

# Write the template for each group horizontally
for idx, group in enumerate(groups):
    start_col = idx * columns_per_group + 3  # Start from column 3 (C)

    # Write the group name and merge cells across 4 columns
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 3)
    group_cell = ws.cell(row=3, column=start_col)
    group_cell.value = group
    group_cell.font = header_font
    group_cell.fill = header_fill
    group_cell.alignment = center_alignment

    # Write headers (Fan, Dars turi, O'qituvchi, Xona)
    ws.cell(row=4, column=start_col).value = "Fan"
    ws.cell(row=4, column=start_col + 1).value = "Dars turi"
    ws.cell(row=4, column=start_col + 2).value = "O'qituvchi"
    ws.cell(row=4, column=start_col + 3).value = "Xona"

    # Apply styles to headers
    for col in range(start_col, start_col + 4):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Apply borders to all cells in the used range
    for row in range(3, 5 + 6 * len(days)):  # Rows 3 to 40 (6 days * 6 periods + headers)
        for col in range(1, (len(groups) * columns_per_group) + 3):  # Columns A to J (or more if groups increase)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

# Save the workbook
wb.save("schedule_template_updated.xlsx")